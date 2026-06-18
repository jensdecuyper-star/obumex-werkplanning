#!/usr/bin/env python3
"""
plaatsing-sync.py
Sync plaatsingsdagen (setup days) from PROGRAMMERING CNC + FileMaker data
Writes to Firestore planning/plaatsing_live
"""

import json
import csv
import os
import re
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
import requests

def isoWeekKey(d):
    """ISO week key: YYYY-Www"""
    jan4 = datetime(d.year, 1, 4)
    wk1Mon = jan4 - timedelta(days=jan4.weekday())
    diffDays = (d - wk1Mon).days
    weekNum = diffDays // 7 + 1
    return f"{d.year}-W{weekNum:02d}"

def extract_firebase_config(index_html_path):
    """Extract Firebase config from index.html"""
    if not os.path.exists(index_html_path):
        return None
    with open(index_html_path, 'r', encoding='utf-8') as f:
        html = f.read()
    m = re.search(r'\{[^{}]*apiKey[^{}]*\}', html)
    if not m:
        return None
    try:
        config = eval('(' + m.group(0) + ')')
        return config
    except:
        return None

def read_programmering_cnc(excel_path):
    """
    Read PROGRAMMERING CNC Excel file
    Return dict: {(project, fase) -> plaatsingsdagen}
    """
    plaatsing_map = {}
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb['Programmeringslijst']

        # Find column indices
        headers = {}
        for col_idx, cell in enumerate(ws[3], 1):
            if cell.value:
                headers[cell.value.strip()] = col_idx

        if 'PLAATSINGSDAGEN' not in headers:
            print('[plaatsing-sync] WARNING: PLAATSINGSDAGEN column not found')
            return plaatsing_map

        plaatsing_col = headers['PLAATSINGSDAGEN']
        project_col = headers.get('PROJECT', 2)
        fase_col = headers.get('FASE', 3)
        startdat_col = headers.get('STARTDATUM PLAATSERS', 12)

        # Read rows
        for row_idx in range(4, ws.max_row + 1):
            project = ws.cell(row_idx, project_col).value
            fase = ws.cell(row_idx, fase_col).value
            plaatsing = ws.cell(row_idx, plaatsing_col).value
            startdat = ws.cell(row_idx, startdat_col).value

            if project and fase and plaatsing:
                try:
                    p = str(project).strip()
                    f = str(int(float(fase))) if isinstance(fase, (int, float)) else str(fase).strip()
                    pdays = float(plaatsing)
                    week_key = None
                    if isinstance(startdat, datetime):
                        week_key = isoWeekKey(startdat)
                    plaatsing_map[(p, f)] = {'days': pdays, 'week': week_key}
                except:
                    pass

        print(f'[plaatsing-sync] Read {len(plaatsing_map)} fases with plaatsingsdagen')
        return plaatsing_map
    except Exception as e:
        print(f'[plaatsing-sync] Error reading Excel: {e}')
        return {}

def read_filemaker_afgewerkt(csv_path):
    """
    Read FileMaker afgewerkt data
    CSV: datum,project,klant,ruimte,doorlooptijd,maand
    Project format: P24-0317-125100 (project+fase combined)
    Return dict: {(project, fase) -> week_key}
    """
    afgewerkt = {}
    try:
        if not os.path.exists(csv_path):
            print(f'[plaatsing-sync] No CSV found: {csv_path}')
            return afgewerkt

        import csv
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    datum_str = row.get('datum', '')
                    proj_fase = row.get('project', '')
                    
                    if not datum_str or not proj_fase:
                        continue
                    
                    d = datetime.strptime(datum_str, '%Y-%m-%d')
                    week = isoWeekKey(d)
                    
                    # Split P24-0317-125100 -> P24-0317, 125100
                    parts = proj_fase.split('-')
                    if len(parts) >= 3:
                        project = '-'.join(parts[:-1])
                        fase = parts[-1]
                        afgewerkt[(project, fase)] = week
                except:
                    pass

        print(f'[plaatsing-sync] Read {len(afgewerkt)} afgewerkte fases')
        return afgewerkt
    except Exception as e:
        print(f'[plaatsing-sync] Error reading CSV: {e}')
        return {}

def calculate_plaatsing_per_week(plaatsing_map, afgewerkt_map):
    """
    Match plaatsingsdagen to afgewerkte weeks
    Return: {week_key -> total_plaatsingsdagen}, detail
    """
    per_week = {}
    per_week_detail = {}

    for (project, fase), pinfo in plaatsing_map.items():
        pdays = pinfo['days']
        if (project, fase) in afgewerkt_map:
            week = afgewerkt_map[(project, fase)]
            per_week[week] = per_week.get(week, 0) + pdays

            if week not in per_week_detail:
                per_week_detail[week] = []
            per_week_detail[week].append({
                'project': project,
                'fase': fase,
                'plaatsingsdagen': pdays
            })

    print(f'[plaatsing-sync] Matched plaatsingsdagen for {len(per_week)} weeks')
    return per_week, per_week_detail

def write_firestore(fb_config, payload):
    """Write to Firestore planning/plaatsing_live"""
    if not fb_config:
        print('[plaatsing-sync] No Firebase config. Skipping Firestore write.')
        return True
    
    try:
        api_key = fb_config.get('apiKey')
        project_id = fb_config.get('projectId')

        url = f"https://firestore.googleapis.com/v1/projects/{project_id}/databases/(default)/documents/planning/plaatsing_live?key={api_key}"

        doc = {
            'fields': {
                'payload': {'stringValue': json.dumps(payload)},
                'updatedAt': {'timestampValue': datetime.utcnow().isoformat() + 'Z'}
            }
        }

        resp = requests.patch(url, json={'fields': doc['fields']}, timeout=10)

        if resp.status_code in (200, 201):
            print('[plaatsing-sync] ✓ Firestore written: planning/plaatsing_live')
            return True
        else:
            print(f'[plaatsing-sync] Firestore error: {resp.status_code}')
            return False
    except Exception as e:
        print(f'[plaatsing-sync] Error writing Firestore: {e}')
        return False

def main():
    print('[plaatsing-sync] Start')

    # Paths
    cwd = Path.cwd()
    index_html = cwd / 'index.html'
    excel_file = cwd / 'programmering_cnc.xlsx'
    csv_file = cwd / '_afgewerkt_clean.csv'

    # 1. Extract Firebase config
    fb_config = extract_firebase_config(str(index_html))
    if not fb_config:
        print('[plaatsing-sync] WARNING: No Firebase config found')

    # 2. Read plaatsingsdagen from Excel
    if not os.path.exists(excel_file):
        print(f'[plaatsing-sync] ERROR: Excel file not found: {excel_file}')
        return

    plaatsing_map = read_programmering_cnc(str(excel_file))
    if not plaatsing_map:
        print('[plaatsing-sync] ERROR: No plaatsingsdagen found in Excel')
        return

    # 3. Read FileMaker afgewerkt data
    afgewerkt_map = read_filemaker_afgewerkt(str(csv_file))

    # 4. Calculate plaatsingsdagen per week
    per_week, per_week_detail = calculate_plaatsing_per_week(plaatsing_map, afgewerkt_map)

    # 5. Build payload
    total_plaatsing = sum(p['days'] for p in plaatsing_map.values())
    payload = {
        'totaalPlaatsingsdagen': round(total_plaatsing, 1),
        'plaatsingPerWeek': {w: round(p, 1) for w, p in sorted(per_week.items())},
        'plaatsingFases': per_week_detail,
        'matchedFases': len(set(afgewerkt_map.keys()) & set(plaatsing_map.keys())),
        'totaleFases': len(plaatsing_map)
    }

    print(f'[plaatsing-sync] Payload: totaal={payload["totaalPlaatsingsdagen"]}d, weeks={len(per_week)}, matched={payload["matchedFases"]}')
    print(f'[plaatsing-sync] Per week: {list(per_week.items())[:3]}...')

    # 6. Write to Firestore
    write_firestore(fb_config, payload)

    print('[plaatsing-sync] Done')

if __name__ == '__main__':
    main()
